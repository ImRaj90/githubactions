import openpyxl
import requests 
from bs4 import BeautifulSoup
wb= openpyxl.load_workbook('Master Microsoft list 1.xlsx')
sheet=wb['Sheet1']
max_row = len(sheet['A'])
print(max_row)
url = "https://www.alexa.com/siteinfo/"
for x in range (1,max_row):
    fullurl = url+sheet.cell(row=x,column=1).value
    #print(x,sheet.cell(row=x,column=1).value)
    try:
        reqs = requests.get(fullurl) 
    except:
        continue    
    # using the BeaitifulSoup module 
    soup = BeautifulSoup(reqs.text, 'html.parser') 
    temp="b"+ str(x)
    #print(x,sheet.cell(row=x,column=1).value, temp, sheet[temp])
    fuck = soup.find_all("p", {"class": "big data"})
    if fuck:
        for title in soup.find_all("p", {"class": "big data"}): 
           # print(title.get_text())
            s = title.get_text()
            s = s.strip()
            s1 = s[1:]
            s1 = s1.strip()
            s1.replace(" ", "") 
            s1 = s1.replace(',', '')
            v1 = int(s1)
            #print("here is the v1",v1)    
            if(v1>1 and v1<=100):
                sheet[temp]=1000000000
               # print(sheet[temp])
            elif(v1>101 and v1<=1000):
                sheet[temp] = 128100000
            elif(v1>1001 and v1<=10000):
                sheet[temp] = 40200000
            elif(v1>10001 and v1<=50000):
                sheet[temp] = 5000000
            elif(v1>50001 and v1<=100000):
                sheet[temp] = 300000
            elif(v1>100001 and v1<=500000):
                sheet[temp] = 150000
            elif(v1>500001 and v1<=1000000):
                sheet[temp] = 15000
            else:
                sheet[temp] = 5000
            #    print("this is from else", sheet[temp])
             #   print("this is from face", s1)
    else:
        s1=0
        sheet[temp]=5000
       # print(sheet[temp], sep=" | ")
    if x % 100 == 0:
        print(x,"is divided by 100",x % 100)   
        wb.save('Master Microsoft list 1.xlsx')           

    print(x,sheet.cell(row=x,column=1).value, temp, sheet[temp].value, sep=" | ")
wb.save('Master Microsoft list 1.xlsx')
